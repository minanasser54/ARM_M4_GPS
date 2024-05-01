
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

def draw_3d_map(longitude, latitude, filename):
    # Get current directory
    current_dir = os.getcwd()

    # Join current directory with the filename
    filepath = os.path.join(current_dir, filename)

    # Create a new workbook
    wb = Workbook()
    ws = wb.active

    # Write headers
    ws['A1'] = "Longitude"
    ws['B1'] = "Latitude"

    # Write data
    for i, (lon, lat) in enumerate(zip(longitude, latitude), start=2):
        ws[f'A{i}'] = lon
        ws[f'B{i}'] = lat

    # Add VBA code to create 3D Map
    vba_code = """
    Sub Create3DMap()
        Dim objTable As ListObject
        Dim objChart As Chart

        ' Convert data to table
        Set objTable = ActiveSheet.ListObjects.Add(xlSrcRange, Range("A1").CurrentRegion, , xlYes)
        
        ' Create 3D Map
        Set objChart = ActiveSheet.Shapes.AddChart2(1, xl3DMapsSurface).Chart
        objChart.SetSourceData Source:=objTable, PlotBy:=xlColumns
        objChart.ChartTitle.Text = "3D Map"
    End Sub
    """

    # Write VBA code to worksheet
    ws.vba_code = vba_code

    # Save the workbook in the same folder
    wb.save(filepath)

    print(f"3D Map data written to {filepath} successfully!")

# Sample longitude and latitude data
longitude = [40.7128, 34.0522, 51.5074]
latitude = [-74.0060, -118.2437, -0.1278]
filename = "3d_map_data.xlsx"

# Draw 3D Map
draw_3d_map(longitude, latitude, filename)